#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Alpaca Crypto (BTC/ETH/...) Compounding Paper Trader V3
========================================================
Runs in Pythonista on iPad and in VS Code on a laptop. Pure stdlib + `requests`;
`openpyxl` is optional (CSV fallback if missing).

What this version does
----------------------
- Trades crypto on Alpaca PAPER (e.g. BTC/USD, ETH/USD), 24/7, no market hours.
- Strategy is informed by Ernest P. Chan, "Quantitative Trading" (2nd ed.):
    * Mean-reversion: buy when price is in the lower portion of a rolling window
      AND below a configurable "buy rate of the day" anchor (daily dip %).
    * Take-profit threshold targeting 1-2% per cycle (compounded across cycles).
    * Trailing stop AND hard stop-loss per position.
    * Kelly-inspired position fraction so position size scales with capital, not
      a fixed dollar amount.
- Compound growth: realized P/L from each closed campaign is added to a
  compounding pool that increases the next buy size.
- Risk caps: hard stop-loss %, daily-loss cap (USD), consecutive-loss circuit
  breaker, total exposure cap.
- THREE emergency stops:
    STOP.txt              -> finish current cycle, save dashboard, exit.
    SELL_ALL_AND_STOP.txt -> liquidate open positions, save, exit.
    PANIC.txt             -> exit immediately. NO orders are placed (bot dies).

IMPORTANT REALITY CHECK
-----------------------
A target of 1-2% per day = 365-730% per year compounded. No automated strategy
reliably delivers that. This bot uses 1-2% as the *take-profit threshold per
trade cycle* - it sells when up that much. It does not guarantee that every
day will be a winning day. Use PAPER trading first. Verify behavior. Read all
the warnings the bot prints.

Files created (in the script's working directory)
-------------------------------------------------
  crypto_bot_state_v3.json      - persistent state across restarts
  crypto_bot_dashboard.xlsx     - live Excel dashboard (if openpyxl installed)
  crypto_trade_log_v3.csv       - every trade, even without openpyxl
  STOP.txt / SELL_ALL_AND_STOP.txt / PANIC.txt - emergency triggers

Usage
-----
  python alpaca_crypto_bot_v3.py            # run forever (default)
  python alpaca_crypto_bot_v3.py --once     # one cycle and exit (cron-friendly)
  python alpaca_crypto_bot_v3.py --status   # print account + position summary
  python alpaca_crypto_bot_v3.py --reset    # wipe state file (keeps dashboard)
"""

import os
import sys
import csv
import json
import time
import urllib.parse
from datetime import datetime, timedelta, timezone

import requests

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except Exception:
    OPENPYXL_OK = False


# =============================================================================
# CONFIG  -  EDIT EVERYTHING IN THIS SECTION TO TUNE THE BOT
# =============================================================================

# ---- API CREDENTIALS --------------------------------------------------------
# Get keys from https://app.alpaca.markets -> Paper Trading -> API Keys.
# Either edit these strings OR set the env vars (env vars take precedence).
ALPACA_API_KEY_ID     = os.getenv("ALPACA_API_KEY_ID",     "PUT_NEW_PAPER_KEY_HERE")
ALPACA_API_SECRET_KEY = os.getenv("ALPACA_API_SECRET_KEY", "PUT_NEW_PAPER_SECRET_HERE")

# Paper trading endpoints. Do NOT change to live until you have weeks of paper
# results that you trust.
BASE_TRADING_URL = "https://paper-api.alpaca.markets"
BASE_DATA_URL    = "https://data.alpaca.markets"

# ---- LOOP TIMING ------------------------------------------------------------
RUN_EVERY_SECONDS  = 120      # how often the main loop ticks (crypto is 24/7)
REQUEST_TIMEOUT    = 20
ORDER_FILL_WAIT_S  = 20       # how long to wait for a market order to fill
ORDER_FILL_POLL_S  = 1.0

# ---- GLOBAL RISK CAPS  (the "no losses" guard rails) ------------------------
# These are HARD caps. The bot will refuse to trade or will liquidate if hit.
MIN_CASH_USD              = 50.0     # never spend below this much cash
MAX_TOTAL_EXPOSURE_USD    = 2000.0   # total $ across all open crypto positions
HARD_STOP_LOSS_PCT        = 0.015    # 1.5% per-position emergency exit (hard)
MAX_DAILY_LOSS_USD        = 50.0     # if today's realized loss >= this, pause
MAX_CONSECUTIVE_LOSSES    = 3        # auto-pause after N losing campaigns in a row
PAUSE_AFTER_CIRCUIT_HOURS = 6        # how long the circuit-breaker keeps you out

# ---- COMPOUNDING -----------------------------------------------------------
# Starting bankroll allocated to the bot. Each closed campaign's realized P/L
# is added to this pool and used to scale the next buy size.
COMPOUNDING_BASE_USD = 500.0

# Fraction of the compounded pool to deploy on a single buy signal.
# Kelly-inspired (Chan ch. 6). Keep this conservative; full Kelly is too hot.
KELLY_FRACTION = 0.20    # use 20% of compound pool per trade (= 5 trades to deploy)

# Hard floor and ceiling on a single buy size, regardless of the pool.
MIN_BUY_NOTIONAL_USD = 25.0
MAX_BUY_NOTIONAL_USD = 500.0

# ---- WATCHLIST (per-symbol settings) ----------------------------------------
# Add or remove symbols here. Alpaca crypto symbols use a slash, e.g. BTC/USD.
# Each block has its own buy rate of the day, sell threshold, etc.
WATCHLIST = [
    {
        "symbol":               "BTC/USD",
        "enabled":              True,
        # ---- BUY RATE OF THE DAY ----
        # The daily anchor is captured at first run of each UTC day. The bot
        # only buys if price has dipped this far BELOW the day's anchor.
        # 0.005 = 0.5% dip from the day's anchor.
        "daily_buy_dip_pct":    0.005,
        # Bar timeframe + lookback window for the rolling-range filter.
        # Valid timeframes: 1Min, 5Min, 15Min, 1Hour, 4Hour, 1Day.
        "timeframe":            "15Min",
        "lookback_hours":       12,
        # On top of the daily-dip rule, also require the price to be in the
        # bottom X% of the recent window (mean-reversion confirmation).
        "buy_range_fraction":   0.30,    # bottom 30% of window
        # ---- SELLING THRESHOLD ----
        # Take profit on the trade. 0.012 = 1.2% gain. Tune to your 1-2%/day goal.
        "profit_target_pct":    0.012,
        # Trailing stop: lock in gains. After price has moved up post-entry,
        # sell if it pulls back this much from the post-entry peak.
        "trailing_stop_pct":    0.008,
        # Optional: also sell if we are near the top of the rolling range.
        "sell_range_fraction":  0.85,
    },
    {
        "symbol":               "ETH/USD",
        "enabled":              False,           # flip to True to enable
        "daily_buy_dip_pct":    0.007,
        "timeframe":            "15Min",
        "lookback_hours":       12,
        "buy_range_fraction":   0.30,
        "profit_target_pct":    0.015,
        "trailing_stop_pct":    0.010,
        "sell_range_fraction":  0.85,
    },
]

# ---- FILES ------------------------------------------------------------------
STATE_FILE             = "crypto_bot_state_v3.json"
DASHBOARD_XLSX         = "crypto_bot_dashboard.xlsx"
TRADE_LOG_CSV          = "crypto_trade_log_v3.csv"
STOP_FILE              = "STOP.txt"
SELL_ALL_AND_STOP_FILE = "SELL_ALL_AND_STOP.txt"
PANIC_FILE             = "PANIC.txt"

# =============================================================================
# END OF CONFIG  -  do not edit below unless you know what you're doing
# =============================================================================


HEADERS = {
    "APCA-API-KEY-ID":     ALPACA_API_KEY_ID,
    "APCA-API-SECRET-KEY": ALPACA_API_SECRET_KEY,
    "Content-Type":        "application/json",
    "Accept":              "application/json",
}


# -----------------------------------------------------------------------------
# BASIC HELPERS
# -----------------------------------------------------------------------------
def now_utc():           return datetime.now(timezone.utc)
def now_iso():           return now_utc().isoformat()
def utc_date_str():      return now_utc().date().isoformat()
def usd(x):              return f"${float(x):,.2f}"
def pct(v):              return f"{float(v) * 100:.2f}%"
def print_status(msg):   print(f"[{now_iso()}] {msg}", flush=True)


def load_state():
    if os.path.exists(STATE_FILE):
        try:
            with open(STATE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            print_status(f"WARN: state file unreadable, starting fresh: {e}")
    return {
        "symbols":              {},
        "closed_campaigns":     [],
        "compounding_pool_usd": float(COMPOUNDING_BASE_USD),
        "consecutive_losses":   0,
        "circuit_breaker_until": "",   # ISO timestamp; "" = no breaker
        "daily_realized_pl":    {},    # {"YYYY-MM-DD": float}
        "daily_anchors":        {},    # {"YYYY-MM-DD": {"BTC/USD": price, ...}}
    }


def save_state(state):
    tmp = STATE_FILE + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(state, f, indent=2)
    os.replace(tmp, STATE_FILE)


def ensure_trade_csv():
    if not os.path.exists(TRADE_LOG_CSV):
        with open(TRADE_LOG_CSV, "w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerow([
                "timestamp_utc", "campaign_id", "symbol", "side",
                "filled_qty", "filled_avg_price", "gross_value_usd",
                "order_id", "reason", "note",
            ])


def append_trade_csv(row):
    ensure_trade_csv()
    with open(TRADE_LOG_CSV, "a", newline="", encoding="utf-8") as f:
        csv.writer(f).writerow(row)


def api_get(url, params=None):
    r = requests.get(url, headers=HEADERS, params=params, timeout=REQUEST_TIMEOUT)
    if r.status_code >= 300:
        raise RuntimeError(f"GET {url} -> {r.status_code} {r.text}")
    return r.json()


def api_post(url, payload):
    r = requests.post(url, headers=HEADERS, json=payload, timeout=REQUEST_TIMEOUT)
    if r.status_code >= 300:
        raise RuntimeError(f"POST {url} -> {r.status_code} {r.text}")
    return r.json()


def url_encode_symbol(symbol):
    """BTC/USD -> BTC%2FUSD (for use inside path segments)."""
    return urllib.parse.quote(symbol, safe="")


def position_symbol_variants(symbol):
    """
    Alpaca's positions endpoint accepts either 'BTC/USD' or 'BTCUSD'. Different
    accounts/regions seem to behave differently, so we accept both shapes.
    """
    return [symbol, symbol.replace("/", "")]


# -----------------------------------------------------------------------------
# ALPACA API  (crypto-specific)
# -----------------------------------------------------------------------------
def get_account():
    return api_get(f"{BASE_TRADING_URL}/v2/account")


def get_position(symbol):
    last_err = None
    for sym in position_symbol_variants(symbol):
        try:
            return api_get(
                f"{BASE_TRADING_URL}/v2/positions/{url_encode_symbol(sym)}"
            )
        except Exception as e:
            if "404" in str(e):
                last_err = e
                continue
            raise
    # 404 on every variant means flat
    return None


def get_latest_crypto_price(symbol):
    """
    Crypto market data (US): /v1beta3/crypto/us/latest/trades?symbols=BTC/USD
    Falls back to latest quote midpoint if trade is missing.
    """
    js = api_get(
        f"{BASE_DATA_URL}/v1beta3/crypto/us/latest/trades",
        params={"symbols": symbol},
    )
    trade = js.get("trades", {}).get(symbol)
    if trade and "p" in trade:
        return float(trade["p"])

    js2 = api_get(
        f"{BASE_DATA_URL}/v1beta3/crypto/us/latest/quotes",
        params={"symbols": symbol},
    )
    q = js2.get("quotes", {}).get(symbol)
    if q and "bp" in q and "ap" in q:
        return (float(q["bp"]) + float(q["ap"])) / 2.0

    raise RuntimeError(f"Could not fetch latest price for {symbol}")


def get_historical_crypto_bars(symbol, timeframe, lookback_hours, limit=1000):
    end_dt   = now_utc()
    start_dt = end_dt - timedelta(hours=int(lookback_hours) + 1)
    params = {
        "symbols":   symbol,
        "timeframe": timeframe,
        "start":     start_dt.isoformat(),
        "end":       end_dt.isoformat(),
        "limit":     limit,
    }
    js = api_get(f"{BASE_DATA_URL}/v1beta3/crypto/us/bars", params=params)
    bars = js.get("bars", {}).get(symbol, [])
    if not bars:
        raise RuntimeError(f"No historical bars returned for {symbol}")
    return bars


def place_market_buy_notional(symbol, notional_usd):
    payload = {
        "symbol":        symbol,
        "side":          "buy",
        "type":          "market",
        "time_in_force": "gtc",        # crypto requires gtc/ioc, not 'day'
        "notional":      str(round(float(notional_usd), 2)),
    }
    return api_post(f"{BASE_TRADING_URL}/v2/orders", payload)


def place_market_sell_qty(symbol, qty):
    payload = {
        "symbol":        symbol,
        "side":          "sell",
        "type":          "market",
        "time_in_force": "gtc",
        "qty":           str(qty),
    }
    return api_post(f"{BASE_TRADING_URL}/v2/orders", payload)


def get_order(order_id):
    return api_get(f"{BASE_TRADING_URL}/v2/orders/{order_id}")


def wait_for_fill(order_id, timeout_sec=ORDER_FILL_WAIT_S):
    deadline = time.time() + timeout_sec
    last = None
    while time.time() < deadline:
        try:
            order = get_order(order_id)
            last = order
            if str(order.get("status", "")).lower() in (
                "filled", "canceled", "expired", "rejected"
            ):
                return order
        except Exception:
            pass
        time.sleep(ORDER_FILL_POLL_S)
    return last or {"id": order_id, "status": "unknown"}


# -----------------------------------------------------------------------------
# STRATEGY  (mean-reversion + daily anchor + Kelly-inspired sizing)
# -----------------------------------------------------------------------------
def window_stats(bars):
    lows  = [float(b["l"]) for b in bars]
    highs = [float(b["h"]) for b in bars]
    closes = [float(b["c"]) for b in bars]
    low, high = min(lows), max(highs)
    return {
        "window_low":   low,
        "window_high":  high,
        "last_close":   closes[-1],
        "range_position": (closes[-1] - low) / max(high - low, 1e-9),
    }


def get_or_set_daily_anchor(state, symbol, current_price):
    """
    The "buying rate of the day". Captured once per UTC day per symbol; used
    as the reference price for the daily-dip buy rule. Re-anchors at UTC
    midnight automatically.
    """
    today = utc_date_str()
    anchors = state.setdefault("daily_anchors", {}).setdefault(today, {})
    if symbol not in anchors:
        anchors[symbol] = float(current_price)
        print_status(f"{symbol}: daily anchor set @ {current_price:.4f}")
    return float(anchors[symbol])


def evaluate_buy_signal(current_price, daily_anchor, stats, cfg):
    """
    Buy if BOTH:
      - price has dipped >= daily_buy_dip_pct below today's anchor, AND
      - price is in the bottom buy_range_fraction of the rolling window.
    """
    dip_from_anchor = (daily_anchor - current_price) / max(daily_anchor, 1e-9)
    range_pos = (current_price - stats["window_low"]) / max(
        stats["window_high"] - stats["window_low"], 1e-9
    )
    big_enough_dip = dip_from_anchor >= float(cfg["daily_buy_dip_pct"])
    in_bottom_zone = range_pos <= float(cfg["buy_range_fraction"])
    return (big_enough_dip and in_bottom_zone), {
        "dip_from_anchor": dip_from_anchor,
        "range_pos":       range_pos,
        "big_enough_dip":  big_enough_dip,
        "in_bottom_zone":  in_bottom_zone,
        "daily_anchor":    daily_anchor,
    }


def evaluate_sell_signal(current_price, avg_entry, peak_since_entry, stats, cfg):
    """
    Sell if ANY of:
      - profit target reached (your 1-2%/day knob)
      - trailing stop hit
      - HARD STOP LOSS hit (loss-cap "no losses" guard rail)
      - price near the top of the rolling window
    """
    gain_pct      = (current_price - avg_entry) / max(avg_entry, 1e-9)
    drawdown_peak = (peak_since_entry - current_price) / max(peak_since_entry, 1e-9)
    range_pos = (current_price - stats["window_low"]) / max(
        stats["window_high"] - stats["window_low"], 1e-9
    )

    hit_profit_target = gain_pct      >= float(cfg["profit_target_pct"])
    hit_trailing_stop = (drawdown_peak >= float(cfg["trailing_stop_pct"])
                         and current_price > avg_entry)
    hit_hard_stop     = gain_pct      <= -float(HARD_STOP_LOSS_PCT)
    in_top_zone       = range_pos     >= float(cfg["sell_range_fraction"])

    should_sell = hit_profit_target or hit_trailing_stop or hit_hard_stop or in_top_zone
    return should_sell, {
        "gain_pct":          gain_pct,
        "drawdown_peak":     drawdown_peak,
        "hit_profit_target": hit_profit_target,
        "hit_trailing_stop": hit_trailing_stop,
        "hit_hard_stop":     hit_hard_stop,
        "in_top_zone":       in_top_zone,
        "range_pos":         range_pos,
    }


def compute_buy_notional(state):
    """
    Kelly-inspired: spend a fixed fraction of the compounded pool, clamped to
    [MIN_BUY_NOTIONAL_USD, MAX_BUY_NOTIONAL_USD].
    """
    pool = float(state.get("compounding_pool_usd", COMPOUNDING_BASE_USD))
    raw  = pool * float(KELLY_FRACTION)
    return max(MIN_BUY_NOTIONAL_USD, min(MAX_BUY_NOTIONAL_USD, raw))


def total_open_exposure_usd():
    total = 0.0
    for cfg in WATCHLIST:
        if not cfg.get("enabled"):
            continue
        try:
            pos = get_position(cfg["symbol"])
        except Exception:
            pos = None
        if pos:
            total += float(pos.get("market_value", 0.0))
    return total


# -----------------------------------------------------------------------------
# CIRCUIT BREAKERS / RISK CAPS
# -----------------------------------------------------------------------------
def daily_realized_pl(state):
    return float(state.get("daily_realized_pl", {}).get(utc_date_str(), 0.0))


def add_to_daily_pl(state, delta):
    bucket = state.setdefault("daily_realized_pl", {})
    today = utc_date_str()
    bucket[today] = float(bucket.get(today, 0.0)) + float(delta)
    # keep only last 30 days
    cutoff = (now_utc().date() - timedelta(days=30)).isoformat()
    for k in list(bucket.keys()):
        if k < cutoff:
            bucket.pop(k, None)


def circuit_breaker_active(state):
    until = state.get("circuit_breaker_until", "")
    if not until:
        return False, ""
    try:
        until_dt = datetime.fromisoformat(until)
    except Exception:
        return False, ""
    if now_utc() < until_dt:
        return True, until
    state["circuit_breaker_until"] = ""
    return False, ""


def trip_circuit_breaker(state, reason):
    until = (now_utc() + timedelta(hours=PAUSE_AFTER_CIRCUIT_HOURS)).isoformat()
    state["circuit_breaker_until"] = until
    print_status(
        f"!! CIRCUIT BREAKER TRIPPED ({reason}). Trading paused until {until}."
    )


def trading_blocked_by_caps(state):
    """Returns (blocked_bool, reason_str). Sells are still allowed if blocked."""
    active, until = circuit_breaker_active(state)
    if active:
        return True, f"circuit_breaker_until={until}"

    losses = state.get("consecutive_losses", 0)
    if losses >= MAX_CONSECUTIVE_LOSSES:
        trip_circuit_breaker(state, f"{losses} consecutive losses")
        return True, f"consecutive_losses={losses}"

    today_pl = daily_realized_pl(state)
    if today_pl <= -abs(MAX_DAILY_LOSS_USD):
        trip_circuit_breaker(state, f"daily loss {usd(today_pl)} hit cap")
        return True, f"daily_loss={usd(today_pl)}"

    return False, ""


# -----------------------------------------------------------------------------
# CAMPAIGN STATE
# -----------------------------------------------------------------------------
def default_symbol_state(symbol):
    return {
        "symbol":                     symbol,
        "campaign_open":              False,
        "campaign_id":                "",
        "first_buy_timestamp":        "",
        "closed_timestamp":           "",
        "highest_price_since_entry":  0.0,
        "realized_campaign_pl":       0.0,
        "total_bought_usd":           0.0,
        "total_sold_usd":             0.0,
        "last_reason":                "",
        "last_update_utc":            "",
        "window_low":                 0.0,
        "window_high":                0.0,
        "last_buy_notional":          0.0,
    }


def get_symbol_state(state, symbol):
    state.setdefault("symbols", {})
    if symbol not in state["symbols"]:
        state["symbols"][symbol] = default_symbol_state(symbol)
    return state["symbols"][symbol]


def open_campaign(sym_state, symbol):
    ts = now_iso()
    sym_state.update({
        "campaign_open":             True,
        "campaign_id":               f"{symbol.replace('/', '')}_{ts.replace(':', '').replace('-', '')}",
        "first_buy_timestamp":       ts,
        "closed_timestamp":          "",
        "highest_price_since_entry": 0.0,
        "realized_campaign_pl":      0.0,
        "total_bought_usd":          0.0,
        "total_sold_usd":            0.0,
        "last_reason":               "campaign_opened",
        "last_update_utc":           ts,
    })


def close_campaign(state, sym_state, symbol, close_reason):
    sym_state["campaign_open"]        = False
    sym_state["closed_timestamp"]     = now_iso()
    realized = round(
        sym_state["total_sold_usd"] - sym_state["total_bought_usd"], 6
    )
    sym_state["realized_campaign_pl"] = realized
    sym_state["last_reason"]          = close_reason
    sym_state["last_update_utc"]      = now_iso()

    # COMPOUNDING: realized P/L feeds the pool that sizes the next buy.
    state["compounding_pool_usd"] = round(
        float(state.get("compounding_pool_usd", COMPOUNDING_BASE_USD)) + realized, 6
    )
    add_to_daily_pl(state, realized)

    # Consecutive-loss circuit-breaker accounting
    if realized < 0:
        state["consecutive_losses"] = int(state.get("consecutive_losses", 0)) + 1
    else:
        state["consecutive_losses"] = 0

    state.setdefault("closed_campaigns", []).append({
        "campaign_id":         sym_state["campaign_id"],
        "symbol":              symbol,
        "first_buy_timestamp": sym_state["first_buy_timestamp"],
        "closed_timestamp":    sym_state["closed_timestamp"],
        "hours_open":          calc_hours_open(
            sym_state["first_buy_timestamp"], sym_state["closed_timestamp"]
        ),
        "total_bought_usd":    round(sym_state["total_bought_usd"], 6),
        "total_sold_usd":      round(sym_state["total_sold_usd"], 6),
        "realized_pl":         realized,
        "close_reason":        close_reason,
        "compounding_pool_after": state["compounding_pool_usd"],
    })

    print_status(
        f"{symbol}: campaign closed | realized={usd(realized)} "
        f"| compounding_pool={usd(state['compounding_pool_usd'])} "
        f"| consec_losses={state['consecutive_losses']} | reason={close_reason}"
    )


def calc_hours_open(start_iso, end_iso=None):
    if not start_iso:
        return 0.0
    try:
        start = datetime.fromisoformat(start_iso)
        end   = datetime.fromisoformat(end_iso) if end_iso else now_utc()
        return round((end - start).total_seconds() / 3600.0, 3)
    except Exception:
        return 0.0


# -----------------------------------------------------------------------------
# WORKBOOK / DASHBOARD
# -----------------------------------------------------------------------------
SHEET_NAMES = ["Dashboard", "Config", "Daily_Tracker", "Trade_Log",
               "Closed_Campaigns", "Notes"]


def _style_header(ws, row_idx=1, fill_color="1F4E78"):
    fill = PatternFill("solid", fgColor=fill_color)
    thin = Side(style="thin", color="D9D9D9")
    for cell in ws[row_idx]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.border = Border(bottom=thin)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _fit_widths(ws):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            v = "" if cell.value is None else str(cell.value)
            widths[cell.column] = max(widths.get(cell.column, 0), len(v) + 2)
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(w, 12), 30)


def _ensure_workbook():
    if not OPENPYXL_OK:
        return None
    if os.path.exists(DASHBOARD_XLSX):
        wb = load_workbook(DASHBOARD_XLSX)
        for n in SHEET_NAMES:
            if n not in wb.sheetnames:
                wb.create_sheet(n)
        return wb

    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    for n in [s for s in SHEET_NAMES if s != "Dashboard"]:
        wb.create_sheet(n)

    # Dashboard layout
    dash = wb["Dashboard"]
    dash["A1"] = "Alpaca Crypto Compounding Paper Trader V3"
    dash["A1"].font = Font(size=14, bold=True)
    labels = [
        ("A3",  "Last Update (UTC)"),
        ("A4",  "Account Cash"),
        ("A5",  "Portfolio Value"),
        ("A6",  "Buying Power"),
        ("A7",  "Open Exposure (bot)"),
        ("A8",  "Open Symbols"),
        ("A9",  "Compounding Pool"),
        ("A10", "Today Realized P/L"),
        ("A11", "Consecutive Losses"),
        ("A12", "Circuit Breaker Until"),
    ]
    for coord, text in labels:
        dash[coord] = text
    dash["A14"] = "Per-Symbol Summary"
    dash["A14"].font = Font(bold=True)

    headers = [
        "Symbol", "Status", "Campaign ID", "First Buy UTC", "Last Price",
        "Daily Anchor", "Qty", "Avg Entry", "Market Value",
        "Unrealized P/L", "Unrealized P/L %", "Realized Campaign P/L",
        "Peak Since Entry", "Window Low", "Window High",
        "Hours Open", "Last Reason",
    ]
    for _ in range(14):
        dash.append([])
    dash.append(headers)
    _style_header(dash, dash.max_row)

    # Config
    cfg_ws = wb["Config"]
    cfg_ws.append(list(WATCHLIST[0].keys()))
    _style_header(cfg_ws, 1)
    for cfg in WATCHLIST:
        cfg_ws.append([cfg.get(k, "") for k in WATCHLIST[0].keys()])

    # Daily tracker
    daily = wb["Daily_Tracker"]
    daily.append([
        "date_utc", "campaign_id", "symbol", "status", "first_buy_timestamp",
        "last_update_utc", "current_price", "daily_anchor", "position_qty",
        "avg_entry_price", "market_value", "unrealized_pl",
        "unrealized_pl_pct", "realized_campaign_pl",
        "highest_price_since_entry", "window_low", "window_high",
        "hours_open", "close_reason",
    ])
    _style_header(daily, 1)

    # Trade log
    trades = wb["Trade_Log"]
    trades.append([
        "timestamp_utc", "campaign_id", "symbol", "side", "filled_qty",
        "filled_avg_price", "gross_value_usd", "order_id", "reason", "note",
    ])
    _style_header(trades, 1)

    # Closed campaigns
    closed = wb["Closed_Campaigns"]
    closed.append([
        "campaign_id", "symbol", "first_buy_timestamp", "closed_timestamp",
        "hours_open", "total_bought_usd", "total_sold_usd", "realized_pl",
        "close_reason", "compounding_pool_after",
    ])
    _style_header(closed, 1)

    notes = wb["Notes"]
    notes["A1"] = "How to control the bot"
    notes["A1"].font = Font(bold=True)
    notes["A3"]  = "STOP.txt              -> graceful stop after current cycle."
    notes["A4"]  = "SELL_ALL_AND_STOP.txt -> liquidate open positions then stop."
    notes["A5"]  = "PANIC.txt             -> kill instantly (no orders placed)."
    notes["A7"]  = "Edit WATCHLIST and risk caps in the script header."
    notes["A8"]  = "Compounding pool grows from realized P/L; reset by deleting state."
    notes["A10"] = "1-2% per day = 365-730% per year. The bot uses the threshold,"
    notes["A11"] = "not as a guarantee. Run on PAPER first."

    for ws in wb.worksheets:
        _fit_widths(ws)
    wb.save(DASHBOARD_XLSX)
    return wb


def _sheet_row_map(ws, key_cols):
    return {
        tuple(ws.cell(r, c).value for c in key_cols): r
        for r in range(2, ws.max_row + 1)
    }


def update_config_sheet(wb):
    ws = wb["Config"]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    keys = list(WATCHLIST[0].keys())
    for cfg in WATCHLIST:
        ws.append([cfg.get(k, "") for k in keys])
    _fit_widths(ws)


def append_trade_log_sheet(wb, row):
    ws = wb["Trade_Log"]
    ws.append(row)
    _fit_widths(ws)


def upsert_daily_tracker_row(wb, row_dict):
    ws = wb["Daily_Tracker"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    key = (row_dict["date_utc"], row_dict["campaign_id"], row_dict["symbol"])
    rmap = _sheet_row_map(ws, [1, 2, 3])
    if key in rmap:
        r = rmap[key]
        for idx, h in enumerate(headers, start=1):
            ws.cell(r, idx).value = row_dict.get(h, "")
    else:
        ws.append([row_dict.get(h, "") for h in headers])
    _fit_widths(ws)


def append_closed_campaign_sheet(wb, row_dict):
    ws = wb["Closed_Campaigns"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    ws.append([row_dict.get(h, "") for h in headers])
    _fit_widths(ws)


def refresh_dashboard_sheet(wb, account, dashboard_rows, state):
    ws = wb["Dashboard"]
    ws["B3"]  = now_iso()
    ws["B4"]  = float(account.get("cash", 0.0))
    ws["B5"]  = float(account.get("portfolio_value", 0.0))
    ws["B6"]  = float(account.get("buying_power", 0.0))
    ws["B7"]  = sum(float(r.get("market_value", 0.0) or 0.0)
                    for r in dashboard_rows if r.get("status") == "IN_POSITION")
    ws["B8"]  = sum(1 for r in dashboard_rows if r.get("status") == "IN_POSITION")
    ws["B9"]  = float(state.get("compounding_pool_usd", COMPOUNDING_BASE_USD))
    ws["B10"] = daily_realized_pl(state)
    ws["B11"] = int(state.get("consecutive_losses", 0))
    ws["B12"] = state.get("circuit_breaker_until", "") or "—"

    start_row = 16   # under the header at row 15
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)
    for row in dashboard_rows:
        ws.append([
            row.get("symbol", ""), row.get("status", ""),
            row.get("campaign_id", ""), row.get("first_buy_timestamp", ""),
            row.get("last_price", ""), row.get("daily_anchor", ""),
            row.get("qty", ""), row.get("avg_entry_price", ""),
            row.get("market_value", ""), row.get("unrealized_pl", ""),
            row.get("unrealized_pl_pct", ""), row.get("realized_campaign_pl", ""),
            row.get("highest_price_since_entry", ""), row.get("window_low", ""),
            row.get("window_high", ""), row.get("hours_open", ""),
            row.get("last_reason", ""),
        ])

    for coord in ("B4", "B5", "B6", "B7", "B9", "B10"):
        ws[coord].number_format = '$#,##0.00;[Red]($#,##0.00)'
    for r in range(start_row, ws.max_row + 1):
        for col in [5, 6, 8, 9, 12, 13, 14, 15]:
            ws.cell(r, col).number_format = '$#,##0.00;[Red]($#,##0.00)'
        ws.cell(r, 11).number_format = '0.00%'
    _fit_widths(ws)


def save_workbook_safe(wb):
    if wb is not None:
        try:
            wb.save(DASHBOARD_XLSX)
        except PermissionError:
            print_status("WARN: dashboard.xlsx is open in Excel; close it to save.")


# -----------------------------------------------------------------------------
# SNAPSHOT BUILDERS
# -----------------------------------------------------------------------------
def snapshot_daily_row(sym_state, symbol, current_price, daily_anchor, position):
    if position:
        qty               = float(position.get("qty", 0.0))
        avg_entry         = float(position.get("avg_entry_price", 0.0))
        market_value      = float(position.get("market_value", 0.0))
        unrealized_pl     = float(position.get("unrealized_pl", 0.0))
        unrealized_pl_pct = float(position.get("unrealized_plpc", 0.0))
        status            = "IN_POSITION"
    else:
        qty = avg_entry = market_value = unrealized_pl = unrealized_pl_pct = 0.0
        status = "FLAT" if not sym_state.get("campaign_open") else "OPENING"

    return {
        "date_utc":                  utc_date_str(),
        "campaign_id":               sym_state.get("campaign_id", ""),
        "symbol":                    symbol,
        "status":                    status,
        "first_buy_timestamp":       sym_state.get("first_buy_timestamp", ""),
        "last_update_utc":           now_iso(),
        "current_price":             round(float(current_price), 6),
        "daily_anchor":              round(float(daily_anchor), 6),
        "position_qty":              round(qty, 8),
        "avg_entry_price":           round(avg_entry, 6),
        "market_value":              round(market_value, 6),
        "unrealized_pl":             round(unrealized_pl, 6),
        "unrealized_pl_pct":         round(unrealized_pl_pct, 6),
        "realized_campaign_pl":     round(float(sym_state.get("realized_campaign_pl", 0.0)), 6),
        "highest_price_since_entry": round(float(sym_state.get("highest_price_since_entry", 0.0)), 6),
        "window_low":                round(float(sym_state.get("window_low", 0.0)), 6),
        "window_high":               round(float(sym_state.get("window_high", 0.0)), 6),
        "hours_open":                calc_hours_open(sym_state.get("first_buy_timestamp", "")),
        "close_reason":              sym_state.get("last_reason", ""),
    }


def dashboard_row(sym_state, symbol, current_price, daily_anchor, position):
    if position:
        qty       = float(position.get("qty", 0.0))
        avg_entry = float(position.get("avg_entry_price", 0.0))
        mv        = float(position.get("market_value", 0.0))
        upl       = float(position.get("unrealized_pl", 0.0))
        upl_pct   = float(position.get("unrealized_plpc", 0.0))
        status    = "IN_POSITION"
    else:
        qty = avg_entry = mv = upl = upl_pct = 0.0
        if sym_state.get("campaign_open"):
            status = "OPEN_WAITING_FILL"
        elif sym_state.get("campaign_id"):
            status = "CLOSED"
        else:
            status = "FLAT"
    return {
        "symbol":                    symbol,
        "status":                    status,
        "campaign_id":               sym_state.get("campaign_id", ""),
        "first_buy_timestamp":       sym_state.get("first_buy_timestamp", ""),
        "last_price":                round(float(current_price), 6),
        "daily_anchor":              round(float(daily_anchor), 6),
        "qty":                       round(qty, 8),
        "avg_entry_price":           round(avg_entry, 6),
        "market_value":              round(mv, 6),
        "unrealized_pl":             round(upl, 6),
        "unrealized_pl_pct":         round(upl_pct, 6),
        "realized_campaign_pl":     round(float(sym_state.get("realized_campaign_pl", 0.0)), 6),
        "highest_price_since_entry": round(float(sym_state.get("highest_price_since_entry", 0.0)), 6),
        "window_low":                round(float(sym_state.get("window_low", 0.0)), 6),
        "window_high":               round(float(sym_state.get("window_high", 0.0)), 6),
        "hours_open":                calc_hours_open(sym_state.get("first_buy_timestamp", "")),
        "last_reason":               sym_state.get("last_reason", ""),
    }


# -----------------------------------------------------------------------------
# ORDER + LOGGING
# -----------------------------------------------------------------------------
def log_trade(wb, sym_state, symbol, side, order, reason, note=""):
    filled_qty       = float(order.get("filled_qty") or 0.0)
    filled_avg_price = float(order.get("filled_avg_price") or 0.0)
    gross            = round(filled_qty * filled_avg_price, 6)

    row = [
        now_iso(), sym_state.get("campaign_id", ""), symbol, side.upper(),
        round(filled_qty, 8), round(filled_avg_price, 6), gross,
        order.get("id", ""), reason, note,
    ]
    append_trade_csv(row)
    if wb is not None:
        append_trade_log_sheet(wb, row)

    if side.lower() == "buy":
        sym_state["total_bought_usd"]   = round(
            float(sym_state.get("total_bought_usd", 0.0)) + gross, 6)
        sym_state["last_buy_notional"]  = gross
    else:
        sym_state["total_sold_usd"]     = round(
            float(sym_state.get("total_sold_usd", 0.0)) + gross, 6)
        sym_state["realized_campaign_pl"] = round(
            sym_state["total_sold_usd"] - sym_state["total_bought_usd"], 6)


# -----------------------------------------------------------------------------
# CYCLE LOGIC (per-symbol)
# -----------------------------------------------------------------------------
def process_symbol(cfg, state, wb, account_cache=None, force_sell=False):
    symbol     = cfg["symbol"]
    sym_state  = get_symbol_state(state, symbol)

    bars       = get_historical_crypto_bars(symbol, cfg["timeframe"], cfg["lookback_hours"])
    stats      = window_stats(bars)
    current_price = get_latest_crypto_price(symbol)
    position   = get_position(symbol)
    anchor     = get_or_set_daily_anchor(state, symbol, current_price)

    sym_state["window_low"]      = stats["window_low"]
    sym_state["window_high"]     = stats["window_high"]
    sym_state["last_update_utc"] = now_iso()

    if position:
        sym_state["highest_price_since_entry"] = max(
            float(sym_state.get("highest_price_since_entry", 0.0)),
            current_price,
        )

    # -- Force liquidation (SELL_ALL_AND_STOP) --
    if force_sell and position:
        qty = float(position.get("qty", 0.0))
        if qty > 0:
            print_status(f"{symbol}: FORCE SELL {qty} (SELL_ALL_AND_STOP)")
            order  = place_market_sell_qty(symbol, qty)
            filled = wait_for_fill(order["id"])
            log_trade(wb, sym_state, symbol, "sell", filled,
                      "manual_force_sell", "SELL_ALL_AND_STOP")
            refreshed = get_position(symbol)
            if not refreshed:
                close_campaign(state, sym_state, symbol, "manual_force_sell")
                if wb is not None:
                    append_closed_campaign_sheet(wb, state["closed_campaigns"][-1])
            return dashboard_row(sym_state, symbol, current_price, anchor, refreshed)
        return dashboard_row(sym_state, symbol, current_price, anchor, position)

    # -- BUY branch --
    if not position:
        blocked, why = trading_blocked_by_caps(state)
        if blocked:
            print_status(f"{symbol}: BUY BLOCKED ({why})")
            return dashboard_row(sym_state, symbol, current_price, anchor, None)

        if not sym_state.get("campaign_open"):
            buy_signal, info = evaluate_buy_signal(current_price, anchor, stats, cfg)
            print_status(
                f"{symbol}: price={current_price:.4f} | anchor={anchor:.4f} "
                f"| dip={info['dip_from_anchor']*100:.3f}% "
                f"(need {cfg['daily_buy_dip_pct']*100:.2f}%) "
                f"| range_pos={info['range_pos']:.3f}"
            )

            acct  = account_cache or get_account()
            cash  = float(acct.get("cash", 0.0))
            buy_notional   = compute_buy_notional(state)
            open_exposure  = total_open_exposure_usd()

            can_buy = (
                buy_signal
                and cash >= (buy_notional + MIN_CASH_USD)
                and (open_exposure + buy_notional) <= MAX_TOTAL_EXPOSURE_USD
            )

            if can_buy:
                open_campaign(sym_state, symbol)
                sym_state["highest_price_since_entry"] = current_price
                print_status(
                    f"{symbol}: BUY -> {usd(buy_notional)} "
                    f"(pool={usd(state['compounding_pool_usd'])})"
                )
                order  = place_market_buy_notional(symbol, buy_notional)
                filled = wait_for_fill(order["id"])
                log_trade(wb, sym_state, symbol, "buy", filled,
                          "buy_dip_below_anchor")
                refreshed = get_position(symbol)
                if wb is not None and sym_state.get("campaign_id"):
                    upsert_daily_tracker_row(
                        wb, snapshot_daily_row(
                            sym_state, symbol, current_price, anchor, refreshed
                        )
                    )
                return dashboard_row(sym_state, symbol, current_price, anchor, refreshed)

        if wb is not None and sym_state.get("campaign_id"):
            upsert_daily_tracker_row(
                wb, snapshot_daily_row(sym_state, symbol, current_price, anchor, None)
            )
        return dashboard_row(sym_state, symbol, current_price, anchor, None)

    # -- SELL branch --
    avg_entry = float(position.get("avg_entry_price", 0.0))
    qty       = float(position.get("qty", 0.0))
    sell_signal, info = evaluate_sell_signal(
        current_price=current_price,
        avg_entry=avg_entry,
        peak_since_entry=float(sym_state.get("highest_price_since_entry", current_price)),
        stats=stats,
        cfg=cfg,
    )
    print_status(
        f"{symbol}: price={current_price:.4f} | entry={avg_entry:.4f} | qty={qty:.8f} "
        f"| gain={info['gain_pct']*100:.2f}% "
        f"| dd_from_peak={info['drawdown_peak']*100:.2f}% "
        f"| range_pos={info['range_pos']:.3f}"
    )

    if sell_signal:
        reasons = []
        if info["hit_profit_target"]: reasons.append("profit_target")
        if info["hit_trailing_stop"]: reasons.append("trailing_stop")
        if info["hit_hard_stop"]:     reasons.append("HARD_STOP_LOSS")
        if info["in_top_zone"]:       reasons.append("top_zone")
        reason = "|".join(reasons) or "sell_signal"
        print_status(f"{symbol}: SELL {qty} ({reason})")

        order  = place_market_sell_qty(symbol, qty)
        filled = wait_for_fill(order["id"])
        log_trade(wb, sym_state, symbol, "sell", filled, reason)
        refreshed = get_position(symbol)
        if not refreshed:
            close_campaign(state, sym_state, symbol, reason)
            if wb is not None:
                append_closed_campaign_sheet(wb, state["closed_campaigns"][-1])
        if wb is not None and sym_state.get("campaign_id"):
            upsert_daily_tracker_row(
                wb, snapshot_daily_row(
                    sym_state, symbol, current_price, anchor, refreshed
                )
            )
        return dashboard_row(sym_state, symbol, current_price, anchor, refreshed)

    # -- HOLD branch --
    if wb is not None and sym_state.get("campaign_id"):
        upsert_daily_tracker_row(
            wb, snapshot_daily_row(sym_state, symbol, current_price, anchor, position)
        )
    return dashboard_row(sym_state, symbol, current_price, anchor, position)


# -----------------------------------------------------------------------------
# WHOLE-CYCLE ORCHESTRATION
# -----------------------------------------------------------------------------
def refresh_all_dashboards(wb, account, dashboard_rows, state):
    if wb is None:
        return
    update_config_sheet(wb)
    refresh_dashboard_sheet(wb, account, dashboard_rows, state)
    save_workbook_safe(wb)


def run_once(force_sell_all=False):
    state   = load_state()
    wb      = _ensure_workbook()
    account = get_account()
    cash    = float(account.get("cash", 0.0))

    print_status(
        f"Account cash={usd(cash)} | portfolio={usd(account.get('portfolio_value', 0.0))} "
        f"| pool={usd(state.get('compounding_pool_usd', COMPOUNDING_BASE_USD))} "
        f"| today_pl={usd(daily_realized_pl(state))}"
    )

    blocked, why = trading_blocked_by_caps(state)
    if blocked and not force_sell_all:
        print_status(f"!! TRADING PAUSED ({why}). Sells still allowed if signaled.")

    dashboard_rows = []
    for cfg in WATCHLIST:
        if not cfg.get("enabled"):
            continue
        try:
            dashboard_rows.append(
                process_symbol(cfg, state, wb,
                               account_cache=account,
                               force_sell=force_sell_all)
            )
        except Exception as e:
            print_status(f"{cfg['symbol']}: ERROR -> {e}")
            sym_state = get_symbol_state(state, cfg["symbol"])
            dashboard_rows.append({
                "symbol":             cfg["symbol"],
                "status":             "ERROR",
                "campaign_id":        sym_state.get("campaign_id", ""),
                "first_buy_timestamp": sym_state.get("first_buy_timestamp", ""),
                "last_price":         "",
                "daily_anchor":       "",
                "qty":                "",
                "avg_entry_price":    "",
                "market_value":       "",
                "unrealized_pl":      "",
                "unrealized_pl_pct":  "",
                "realized_campaign_pl": round(
                    float(sym_state.get("realized_campaign_pl", 0.0)), 6),
                "highest_price_since_entry": round(
                    float(sym_state.get("highest_price_since_entry", 0.0)), 6),
                "window_low":         round(float(sym_state.get("window_low", 0.0)), 6),
                "window_high":        round(float(sym_state.get("window_high", 0.0)), 6),
                "hours_open":         calc_hours_open(sym_state.get("first_buy_timestamp", "")),
                "last_reason":        str(e),
            })

    refresh_all_dashboards(wb, get_account(), dashboard_rows, state)
    save_state(state)


def finalize_open_campaign_rows(state, wb):
    """Called on graceful stop so the dashboard reflects the last snapshot."""
    rows = []
    for cfg in WATCHLIST:
        if not cfg.get("enabled"):
            continue
        symbol = cfg["symbol"]
        sym_state = get_symbol_state(state, symbol)
        try:
            current_price = get_latest_crypto_price(symbol)
            position      = get_position(symbol)
            anchor        = get_or_set_daily_anchor(state, symbol, current_price)
            if sym_state.get("campaign_id"):
                row = snapshot_daily_row(sym_state, symbol, current_price, anchor, position)
                row["status"] = "MONITOR_STOPPED" if position else "CLOSED"
                if wb is not None:
                    upsert_daily_tracker_row(wb, row)
            rows.append(dashboard_row(sym_state, symbol, current_price, anchor, position))
        except Exception as e:
            print_status(f"{symbol}: finalize error: {e}")
    try:
        refresh_all_dashboards(wb, get_account(), rows, state)
    except Exception as e:
        print_status(f"Dashboard finalize error: {e}")


# -----------------------------------------------------------------------------
# CLI MODES
# -----------------------------------------------------------------------------
def cmd_status():
    state   = load_state()
    account = get_account()
    print(f"Account cash:        {usd(account.get('cash', 0.0))}")
    print(f"Portfolio value:     {usd(account.get('portfolio_value', 0.0))}")
    print(f"Buying power:        {usd(account.get('buying_power', 0.0))}")
    print(f"Compounding pool:    {usd(state.get('compounding_pool_usd', COMPOUNDING_BASE_USD))}")
    print(f"Today realized P/L:  {usd(daily_realized_pl(state))}")
    print(f"Consecutive losses:  {state.get('consecutive_losses', 0)}")
    print(f"Circuit breaker:     {state.get('circuit_breaker_until', '') or '—'}")
    print()
    for cfg in WATCHLIST:
        sym = cfg["symbol"]
        if not cfg.get("enabled"):
            print(f"  {sym:10} DISABLED")
            continue
        try:
            pos = get_position(sym)
            if pos:
                print(f"  {sym:10} qty={pos.get('qty')} entry={pos.get('avg_entry_price')} "
                      f"upl={pos.get('unrealized_pl')}")
            else:
                print(f"  {sym:10} FLAT")
        except Exception as e:
            print(f"  {sym:10} ERROR: {e}")


def cmd_reset():
    if os.path.exists(STATE_FILE):
        bak = STATE_FILE + ".bak." + now_utc().strftime("%Y%m%d%H%M%S")
        os.replace(STATE_FILE, bak)
        print(f"State moved to {bak}. Dashboard preserved.")
    else:
        print("No state file to reset.")


# -----------------------------------------------------------------------------
# MAIN
# -----------------------------------------------------------------------------
def main():
    if "PUT_NEW_PAPER_KEY_HERE" in ALPACA_API_KEY_ID or \
       "PUT_NEW_PAPER_SECRET_HERE" in ALPACA_API_SECRET_KEY:
        print("ERROR: Set Alpaca PAPER keys (in env vars or in the script) before running.")
        sys.exit(1)

    if "--status" in sys.argv:
        cmd_status()
        return
    if "--reset" in sys.argv:
        cmd_reset()
        return

    if not OPENPYXL_OK:
        print_status("WARNING: openpyxl not installed -> CSV-only logging "
                     "(dashboard.xlsx disabled).")

    if "--once" in sys.argv:
        force_sell = os.path.exists(SELL_ALL_AND_STOP_FILE)
        run_once(force_sell_all=force_sell)
        if force_sell and os.path.exists(SELL_ALL_AND_STOP_FILE):
            os.remove(SELL_ALL_AND_STOP_FILE)
        return

    print_status("Starting Alpaca Crypto Compounding Trader V3 (PAPER)")
    print_status(
        f"Enabled: {', '.join(c['symbol'] for c in WATCHLIST if c.get('enabled')) or '(none)'}"
    )
    print_status(f"Stops: {STOP_FILE} | {SELL_ALL_AND_STOP_FILE} | {PANIC_FILE}")
    print_status(
        f"Caps: hard_stop={pct(HARD_STOP_LOSS_PCT)} | "
        f"daily_loss={usd(MAX_DAILY_LOSS_USD)} | "
        f"max_exposure={usd(MAX_TOTAL_EXPOSURE_USD)} | "
        f"max_consec_losses={MAX_CONSECUTIVE_LOSSES}"
    )

    while True:
        try:
            # PANIC = drop everything immediately, no orders, no save.
            if os.path.exists(PANIC_FILE):
                print_status("!!! PANIC.txt detected. Hard exit, no orders placed. !!!")
                try:
                    os.remove(PANIC_FILE)
                except Exception:
                    pass
                sys.exit(2)

            if os.path.exists(SELL_ALL_AND_STOP_FILE):
                print_status("SELL_ALL_AND_STOP detected.")
                run_once(force_sell_all=True)
                for f in (SELL_ALL_AND_STOP_FILE, STOP_FILE):
                    if os.path.exists(f):
                        try: os.remove(f)
                        except Exception: pass
                break

            if os.path.exists(STOP_FILE):
                print_status("STOP detected. Finalizing and exiting.")
                state = load_state()
                wb    = _ensure_workbook()
                finalize_open_campaign_rows(state, wb)
                save_state(state)
                try: os.remove(STOP_FILE)
                except Exception: pass
                break

            run_once(force_sell_all=False)

        except KeyboardInterrupt:
            print_status("KeyboardInterrupt -> finalizing dashboard.")
            state = load_state()
            wb    = _ensure_workbook()
            finalize_open_campaign_rows(state, wb)
            save_state(state)
            break
        except Exception as e:
            print_status(f"MAIN LOOP ERROR: {e}")

        time.sleep(max(10, int(RUN_EVERY_SECONDS)))


if __name__ == "__main__":
    main()
